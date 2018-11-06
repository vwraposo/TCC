##################################################################################
##                                                                              ##
##  Module that insert the peptide data obtained from a de Novo MS/MS           ##
##                                                                              ##
##################################################################################
import psycopg2
import sys
import os
import re
from openpyxl import load_workbook

try:
    conn = psycopg2.connect(dbname="testdb",  user="fox", password="senha")
except:
    print("Error: It was not possible to connect to the database")
    sys.exit(1)

cur = conn.cursor()

filename = "../data/de_novo.xlsx" 

print("WORKBOOK STARTED")
wb = load_workbook(filename = filename)
print("WORKBOOK LOADED")
sheet = wb.worksheets[0]
print("WORKSHEET LOADED")

for row in sheet.iter_rows(min_row=2):
    data = [cell.value for cell in row]
    species = data[0][2:].strip()

    peptide = data[3].upper()
    #Filter
    peptide = re.sub(r"\(\+[0-9.]*\)", "", peptide)
    if (peptide == ''): 
        continue;
    try:
        cur.execute("SELECT pep_id FROM peptides WHERE pep_seq = '{0}';".format(peptide))
        if (cur.rowcount == 0):
            cur.execute("INSERT INTO peptides(pep_seq, pep_T) VALUES ('{0}', 1)\
                    ON CONFLICT DO NOTHING RETURNING pep_id;".format(peptide))
        pid = cur.fetchone()[0]
        cur.execute("INSERT INTO pep_sn(pep_id, sn_sp) VALUES ('{0}', '{1}')\
                ON CONFLICT DO NOTHING;".format(pid, species))
    except psycopg2.ProgrammingError as e:
        print("Insert error")
        print(e)
        conn.rollback()
        print("Rollback complete")

    conn.commit()


cur.close()
conn.close()

